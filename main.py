from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from serpapi import GoogleSearch
import openpyxl
import time

PATH = 'C:\Program Files (x86)\msedgedriver.exe'

options = Options()
options.add_argument("start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
s = Service(executable_path=PATH)
driver = webdriver.Edge(service=s, options=options)
driver.get('data:,')
driver.get('https://www.amazon.com/')
time.sleep(3)
wb = openpyxl.load_workbook('gamelistformat.xlsx', read_only=False, keep_vba=True)
sheets = wb.sheetnames
ws = wb[sheets[0]]

def target_search(game,platform):
    try:
        game = game.replace(" ","+")
        platform = platform.replace(" ","+")
        driver.get("https://www.target.com/s?searchTerm={}+{}+game".format(game,platform))
        item_web = "/html[1]/body[1]/div[1]/div[2]/main[1]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/section[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]"
        product_div = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, item_web)))
        driver.execute_script("arguments[0].click();", product_div.find_element(By.TAG_NAME,"a"))
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#pdp-product-title-id')))
        price = float(driver.find_element(By.CSS_SELECTOR,".styles__CurrentPriceFontSize-sc-1mh0sjm-1.kwKAiv").text[1:])
        store_title = driver.find_element(By.CSS_SELECTOR, "#pdp-product-title-id").text
        hyperlink = driver.current_url
        driver.get('data:,')
        return store_title, price, hyperlink
    except:
        return 'NA','NA','NA'

def bestbuy_search(game,platform):
    try:
        driver.get('https://www.bestbuy.com/')
        search_bar = driver.find_element(By.XPATH,"//input[@id='gh-search-input']")
        search_bar.send_keys(f"{game} {platform} physical")
        search_bar.send_keys(Keys.RETURN)
        
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div:nth-child(6) > main:nth-child(2) > div:nth-child(4) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(5) > div:nth-child(1) > div:nth-child(4) > ol:nth-child(2) > li:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > h4:nth-child(2) > a:nth-child(1)")))
        item = driver.find_element(By.CSS_SELECTOR,"body > div:nth-child(6) > main:nth-child(2) > div:nth-child(4) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(5) > div:nth-child(1) > div:nth-child(4) > ol:nth-child(2) > li:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > h4:nth-child(2) > a:nth-child(1)")
        driver.execute_script("arguments[0].click();", item)

        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".heading-5.v-fw-regular")))
        price = float(driver.find_element(By.CSS_SELECTOR,"div[class='priceView-hero-price priceView-customer-price'] span[aria-hidden='true']").text[1:])
        store_title = driver.find_element(By.CSS_SELECTOR,".heading-5.v-fw-regular").text
        hyperlink = driver.current_url
        driver.get('data:,')
        return store_title, price, hyperlink
    except:
        return 'NA','NA','NA'

def walmart_search(game,platform):
    return 'test','tes1','test2'
'''
    product_query = game + " " + platform + " game"
    product_query = product_query.replace(" ",'+')
    product_search = "https://www.walmart.com/search?q={}".format(product_query)
    driver.get(product_search)
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//body/div/div/div[@data-testid='layout-container']/div[@aria-hidden='false']/div/div[@data-testid='maincontent']/main/div[@data-testid='flex-container']/div/div[@data-testid='flex-container']/div/div/div/div/section/div/div[1]/div[1]/div[1]/a[1]")))
    store_view = driver.find_element(By.XPATH, "//body/div/div/div[@data-testid='layout-container']/div[@aria-hidden='false']/div/div[@data-testid='maincontent']/main/div[@data-testid='flex-container']/div/div[@data-testid='flex-container']/div/div/div/div/section/div/div[1]/div[1]/div[1]/a[1]")
    product_id_s = store_view.get_attribute("link-identifier")
    params = {
        "engine": "walmart_product",
        "product_id": product_id_s,
        "api_key": "******"}
    search = GoogleSearch(params)
    results = search.get_dict()
    product_result = results["product_result"]
    store_title = product_result['title']
    price = float(product_result['price_map']['price'])
    hyperlink = product_result['product_page_url']
    time.sleep(random.randint(3,13))
    driver.get('data:,')
    return store_title, price, hyperlink
'''
def amazon_search(game,platform):
    try:
        product_query = game + " " + platform + " game"
        product_query = product_query.replace(" ",'+')
        driver.get('https://www.amazon.com/s?k={}'.format(product_query))
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[class='s-widget-container s-spacing-small s-widget-container-height-small celwidget slot=MAIN template=SEARCH_RESULTS widgetId=search-results_1'] span[class='a-size-medium a-color-base a-text-normal']")))
        list_items = driver.find_elements(By.XPATH,"//div[@class = 'sg-col-20-of-24 s-result-item s-asin sg-col-0-of-12 sg-col-16-of-20 sg-col s-widget-spacing-small sg-col-12-of-16']")
        list_items[0].find_element(By.TAG_NAME,"a").click()
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, "//span[@id='productTitle']")))
        store_title = driver.find_element(By.XPATH,"//span[@id='productTitle']").text
        try:
            price_1 = driver.find_element(By.XPATH,"//div[@id='apex_desktop_newAccordionRow']//span[@class='a-price-whole']").text.strip()
            price_2 = driver.find_element(By.CSS_SELECTOR, "div[id='apex_desktop_newAccordionRow'] span[class='a-price-fraction']").text.strip()
        except:
            price_1 = driver.find_element(By.XPATH,"//span[@class='a-price aok-align-center reinventPricePriceToPayMargin priceToPay']//span[@class='a-price-whole']").text.strip()
            price_2 = driver.find_element(By.CSS_SELECTOR,"span[class='a-price aok-align-center reinventPricePriceToPayMargin priceToPay'] span[class='a-price-fraction']").text.strip()
        price = float(price_1 + '.' + price_2)
        hyperlink = driver.current_url
        driver.get('data:,')
        return store_title, price, hyperlink
    except:
        return 'NA','NA','NA'

for row in ws.iter_rows(min_row=4):
    row_check = (row[0].coordinate)
    for i in range(1,5+1):
        if row[i].value != None:
            if i == 1:
                game_console = 'nintendo switch'
                title, price, hyperlink = amazon_search(str(row[0].value),game_console)
                ws['H{}'.format(str(row_check[-1]))] = str(title)
                ws['I{}'.format(str(row_check[-1]))] = str(price)
                ws['J{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = bestbuy_search(str(row[0].value),game_console)
                ws['W{}'.format(str(row_check[-1]))] = str(title)
                ws['X{}'.format(str(row_check[-1]))] = str(price)
                ws['Y{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = target_search(str(row[0].value),game_console)
                ws['AL{}'.format(str(row_check[-1]))] = str(title)
                ws['AM{}'.format(str(row_check[-1]))] = str(price)
                ws['AN{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = walmart_search(str(row[0].value),game_console)
                ws['BA{}'.format(str(row_check[-1]))] = str(title)
                ws['BB{}'.format(str(row_check[-1]))] = str(price)
                ws['BC{}'.format(str(row_check[-1]))] = str(hyperlink)
                pass
            elif i == 2:
                game = 'ps4'
                title, price, hyperlink = amazon_search(str(row[0].value),game_console)
                ws['K{}'.format(str(row_check[-1]))] = str(title)
                ws['L{}'.format(str(row_check[-1]))] = str(price)
                ws['M{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = bestbuy_search(str(row[0].value),game_console)
                ws['Z{}'.format(str(row_check[-1]))] = str(title)
                ws['AA{}'.format(str(row_check[-1]))] = str(price)
                ws['AB{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = target_search(str(row[0].value),game_console)
                ws['AO{}'.format(str(row_check[-1]))] = str(title)
                ws['AP{}'.format(str(row_check[-1]))] = str(price)
                ws['AQ{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = walmart_search(str(row[0].value),game_console)
                ws['BD{}'.format(str(row_check[-1]))] = str(title)
                ws['BE{}'.format(str(row_check[-1]))] = str(price)
                ws['BF{}'.format(str(row_check[-1]))] = str(hyperlink)                
                pass
            elif i == 3:
                game_console = 'ps5'
                title, price, hyperlink = amazon_search(str(row[0].value),game_console)
                ws['N{}'.format(str(row_check[-1]))] = str(title)
                ws['O{}'.format (str(row_check[-1]))] = str(price)
                ws['P{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = bestbuy_search(str(row[0].value),game_console)
                ws['AC{}'.format(str(row_check[-1]))] = str(title)
                ws['AD{}'.format(str(row_check[-1]))] = str(price)
                ws['AE{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = target_search(str(row[0].value),game_console)
                ws['AR{}'.format(str(row_check[-1]))] = str(title)
                ws['AS{}'.format(str(row_check[-1]))] = str(price)
                ws['AT{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = walmart_search(str(row[0].value),game_console)
                ws['BG{}'.format(str(row_check[-1]))] = str(title)
                ws['BH{}'.format(str(row_check[-1]))] = str(price)
                ws['BI{}'.format(str(row_check[-1]))] = str(hyperlink)
                pass
            elif i == 4:
                game_console = 'xbox one'
                title, price, hyperlink = amazon_search(str(row[0].value),game_console)
                ws['Q{}'.format(str(row_check[-1]))] = str(title)
                ws['R{}'.format (str(row_check[-1]))] = str(price)
                ws['S{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = bestbuy_search(str(row[0].value),game_console)
                ws['AF{}'.format(str(row_check[-1]))] = str(title)
                ws['AG{}'.format(str(row_check[-1]))] = str(price)
                ws['AH{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = target_search(str(row[0].value),game_console)
                ws['AU{}'.format(str(row_check[-1]))] = str(title)
                ws['AV{}'.format(str(row_check[-1]))] = str(price)
                ws['AW{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = walmart_search(str(row[0].value),game_console)
                ws['BJ{}'.format(str(row_check[-1]))] = str(title)
                ws['BK{}'.format(str(row_check[-1]))] = str(price)
                ws['BL{}'.format(str(row_check[-1]))] = str(hyperlink)
                pass
            elif i == 5:
                game_console = 'xbox series x'
                title, price, hyperlink = amazon_search(str(row[0].value),game_console)
                ws['T{}'.format(str(row_check[-1]))] = str(title)
                ws['U{}'.format (str(row_check[-1]))] = str(price)
                ws['V{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = bestbuy_search(str(row[0].value),game_console)
                ws['AI{}'.format(str(row_check[-1]))] = str(title)
                ws['AJ{}'.format(str(row_check[-1]))] = str(price)
                ws['AK{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = target_search(str(row[0].value),game_console)
                ws['AX{}'.format(str(row_check[-1]))] = str(title)
                ws['AY{}'.format(str(row_check[-1]))] = str(price)
                ws['AZ{}'.format(str(row_check[-1]))] = str(hyperlink)
                title, price, hyperlink = walmart_search(str(row[0].value),game_console)
                ws['BM{}'.format(str(row_check[-1]))] = str(title)
                ws['BN{}'.format(str(row_check[-1]))] = str(price)
                ws['BO{}'.format(str(row_check[-1]))] = str(hyperlink)                
                pass


wb.save('processed games list.xlsm')
driver.quit()
