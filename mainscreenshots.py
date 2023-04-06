from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
import openpyxl
from datetime import datetime
import os

PATH = 'C:\Program Files (x86)\msedgedriver.exe'

options = Options()
options.add_argument("start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")
s = Service(executable_path=PATH)
driver = webdriver.Edge(service=s, options=options)
wb = openpyxl.load_workbook('gamelistformat.xlsx', read_only=False, keep_vba=True)
sheets = wb.sheetnames
ws = wb[sheets[1]]
now = datetime.now()

try:
    os.makedirs('./screenshots')
except:
    pass

for row in ws.iter_rows(min_row=2):
    if 1 == 1:
        driver.get(row[2].value)
        date_c = now.strftime("%m/%d/%Y")
        name = row[0].value + ' ' + date_c
        name = name.replace('/','')
        print(name)
        driver.save_screenshot('screenshots/{}.png'.format(name))

driver.quit()
